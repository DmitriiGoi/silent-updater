from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from silent_updater.inputs.excel_loader import ExcelFormatError, _cell, _norm_header
from silent_updater.models import Severity, VulnEntry


# Veracode column names we care about. Matched case-insensitively.
COMPONENT_GAV_COL = "Component name and version"
COMPONENT_NAME_COL = "Component name"
VERSION_COL = "Version"
CVE_COL = "Source Ref"
SEVERITY_COL = "Overall Severity"
FIXED_VERSION_COL = "Fixed Version"
SUMMARY_COL = "CVE Summary"


def is_veracode_format(headers: list[str]) -> bool:
    """Heuristic: header set contains Veracode-specific column names."""
    normalized = {h.lower().strip() for h in headers if h}
    return (
        COMPONENT_GAV_COL.lower() in normalized
        and CVE_COL.lower() in normalized
    )


def load_veracode(xlsx_path: str | Path) -> list[VulnEntry]:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ExcelFormatError("Workbook has no active sheet")

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelFormatError("Empty sheet")

    headers = [_norm_header(h) for h in header_row]
    col_index = _build_col_index(headers)

    if "gav" not in col_index:
        raise ExcelFormatError(
            f"Veracode format requires '{COMPONENT_GAV_COL}' column. "
            f"Found: {[h for h in headers if h]}"
        )

    by_key: dict[tuple[str, str], VulnEntry] = {}
    for row_num, row in enumerate(rows, start=2):
        if row is None:
            continue
        gav = _cell(row, col_index["gav"])
        if not gav:
            continue
        try:
            group_id, artifact_id, version = _parse_gav(gav)
        except ValueError:
            continue
        # Prefer explicit Version column if present and non-empty,
        # else fall back to the version inside GAV.
        if "version" in col_index:
            v_explicit = _cell(row, col_index["version"])
            if v_explicit:
                version = v_explicit
        cve = _cell(row, col_index.get("cve", -1))
        severity = Severity.parse(_cell(row, col_index.get("severity", -1)))
        summary = _cell(row, col_index.get("summary", -1))
        fixed_raw = _cell(row, col_index.get("fixed", -1))
        fixed_versions = _extract_versions(fixed_raw) if fixed_raw else ()

        key = (f"{group_id}:{artifact_id}", version)
        if key in by_key:
            existing = by_key[key]
            merged_cves = ", ".join(filter(None, dict.fromkeys(
                [*existing.cve.split(", "), cve]
            ).keys())).strip(", ")
            merged_fix = tuple(dict.fromkeys([*existing.fixed_versions, *fixed_versions]))
            best_sev = existing.severity if existing.severity.rank <= severity.rank else severity
            by_key[key] = VulnEntry(
                group_id=group_id,
                artifact_id=artifact_id,
                vuln_version=version,
                cve=merged_cves,
                severity=best_sev,
                notes=existing.notes or summary,
                fixed_versions=merged_fix,
            )
            continue

        by_key[key] = VulnEntry(
            group_id=group_id,
            artifact_id=artifact_id,
            vuln_version=version,
            cve=cve,
            severity=severity,
            notes=summary,
            fixed_versions=fixed_versions,
        )

    return list(by_key.values())


def _build_col_index(headers: list[str]) -> dict[str, int]:
    lookup = {h.lower().strip(): i for i, h in enumerate(headers) if h}
    idx: dict[str, int] = {}
    if COMPONENT_GAV_COL.lower() in lookup:
        idx["gav"] = lookup[COMPONENT_GAV_COL.lower()]
    if VERSION_COL.lower() in lookup:
        idx["version"] = lookup[VERSION_COL.lower()]
    if CVE_COL.lower() in lookup:
        idx["cve"] = lookup[CVE_COL.lower()]
    if SEVERITY_COL.lower() in lookup:
        idx["severity"] = lookup[SEVERITY_COL.lower()]
    if FIXED_VERSION_COL.lower() in lookup:
        idx["fixed"] = lookup[FIXED_VERSION_COL.lower()]
    if SUMMARY_COL.lower() in lookup:
        idx["summary"] = lookup[SUMMARY_COL.lower()]
    return idx


def _parse_gav(gav: str) -> tuple[str, str, str]:
    """Parse 'groupId:artifactId:version' (or with extra classifier parts) into a 3-tuple.

    Maven GAV can have 3-5 colon-separated parts. We treat the first as groupId,
    the second as artifactId, and the last as version. Anything in between is
    classifier/packaging which we ignore.
    """
    parts = [p.strip() for p in gav.split(":")]
    if len(parts) < 3:
        raise ValueError(f"Not a GAV: {gav!r}")
    return parts[0], parts[1], parts[-1]


_VERSION_TOKEN_RE = re.compile(r"\b\d+(?:\.\d+){1,3}(?:[.\-+][\w.]+)?\b")


def _extract_versions(blob: str) -> tuple[str, ...]:
    """Pull every version-like token out of a Veracode 'Fixed Version' cell.

    Example input: '5.19.3, 5.19.3 ? Version ? 5.19.3, 6.2.2, 6.2.2 ? Version ? 6.2.2'
    Returns: ('5.19.3', '6.2.2') — deduped, original order preserved.
    """
    seen: dict[str, None] = {}
    for match in _VERSION_TOKEN_RE.findall(blob):
        seen.setdefault(match, None)
    return tuple(seen.keys())
