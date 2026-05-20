from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from silent_updater.models import Severity, VulnEntry


REQUIRED_COLUMNS = ("groupId", "artifactId", "vulnerableVersion")
OPTIONAL_COLUMNS = ("cve", "severity", "notes")


class ExcelFormatError(ValueError):
    pass


def load_vulnerable_libs(xlsx_path: str | Path) -> list[VulnEntry]:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ExcelFormatError("Workbook has no active sheet")

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelFormatError("Empty sheet")

    headers = [_norm_header(h) for h in header_row]
    col_index: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h:
            col_index[h] = idx

    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        raise ExcelFormatError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Found headers: {[h for h in headers if h]}"
        )

    entries: list[VulnEntry] = []
    for row_num, row in enumerate(rows, start=2):
        if row is None or all(_is_blank(c) for c in row):
            continue
        group_id = _cell(row, col_index["groupId"])
        artifact_id = _cell(row, col_index["artifactId"])
        vuln_version = _cell(row, col_index["vulnerableVersion"])
        if not (group_id and artifact_id and vuln_version):
            raise ExcelFormatError(
                f"Row {row_num}: groupId/artifactId/vulnerableVersion are required"
            )
        cve = _cell(row, col_index.get("cve", -1))
        severity = Severity.parse(_cell(row, col_index.get("severity", -1)))
        notes = _cell(row, col_index.get("notes", -1))
        entries.append(
            VulnEntry(
                group_id=group_id,
                artifact_id=artifact_id,
                vuln_version=vuln_version,
                cve=cve,
                severity=severity,
                notes=notes,
            )
        )
    return entries


def _norm_header(raw) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def _cell(row, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _is_blank(val) -> bool:
    return val is None or (isinstance(val, str) and not val.strip())
